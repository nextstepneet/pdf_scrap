import os
import io
import re
import json
import ssl
from datetime import datetime
import threading
import uuid

from flask import Flask, request, send_file, render_template
from flask_cors import CORS
from werkzeug.exceptions import HTTPException
from pymongo import MongoClient
from bson import ObjectId
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Fast extractor (pypdfium2)
from extractor import extract_cutoffs, sort_categories

# ─────────────────────────────────────────────────
# App Setup
# ─────────────────────────────────────────────────
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024   # 50 MB max upload
app.config["JSON_AS_ASCII"] = False
CORS(app)

MONGO_URI = os.getenv(
    "MONGODB_URI", 
    "mongodb+srv://shaileshx006067_db_user:qmTowJMpLK063z7K@cluster0.gdh76cd.mongodb.net/?appName=Cluster0"
)

# Lazy MongoDB connection
_mongo_client = None
_col = None

def _make_ssl_ctx():
    """Custom SSL context compatible with Python 3.14 + MongoDB Atlas."""
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    # Allow all TLS versions so we don't get TLSV1_ALERT_INTERNAL_ERROR
    ctx.options &= ~ssl.OP_NO_TLSv1
    ctx.options &= ~ssl.OP_NO_TLSv1_1
    return ctx

def get_col():
    global _mongo_client, _col
    if _col is None:
        try:
            import certifi
            ca_file = certifi.where()
        except ImportError:
            ca_file = None

        try:
            kwargs = {
                "serverSelectionTimeoutMS": 15000,
                "connectTimeoutMS":         15000,
                "socketTimeoutMS":          15000,
            }
            if ca_file:
                kwargs["tls"]       = True
                kwargs["tlsCAFile"] = ca_file
            else:
                kwargs["tls"]                      = True
                kwargs["tlsAllowInvalidCertificates"] = True
                kwargs["tlsAllowInvalidHostnames"]    = True

            _mongo_client = MongoClient(MONGO_URI, **kwargs)
            _mongo_client.admin.command("ping")
            _col = _mongo_client["nextstep_neet"]["extractions"]
            print("[MongoDB] Connected successfully")
        except Exception as e:
            print(f"[MongoDB] Connection failed: {e}")
            _col = None
            raise
    return _col


UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ─────────────────────────────────────────────────
# Excel Generation
# ─────────────────────────────────────────────────
def generate_excel(records: list) -> io.BytesIO:
    """
    Rows    = Colleges (code + name)
    Columns = Categories
    Value   = Last (highest AIR) cutoff rank
    """
    colleges = sorted(records, key=lambda r: r["college_code"])

    all_cats: set = set()
    for r in records:
        all_cats.update(r["category_cutoffs"].keys())
    categories = sort_categories(all_cats)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cutoff Data"

    # ── Styles ──────────────────────────────────────────────────
    title_fill = PatternFill("solid", fgColor="F8FAFC")
    h_fill     = PatternFill("solid", fgColor="E2E8F0")
    even_fill  = PatternFill("solid", fgColor="FFFFFF")
    odd_fill   = PatternFill("solid", fgColor="F8FAFC")
    h_font     = Font(bold=True, color="1E293B", size=10)
    title_font = Font(bold=True, color="0F172A", size=14)
    col_font   = Font(bold=True, color="0F172A", size=10)
    data_font  = Font(size=9, color="334155")
    center_a   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_a     = Alignment(horizontal="left",   vertical="center")
    thin       = Border(
        left  =Side(style="thin", color="CBD5E1"),
        right =Side(style="thin", color="CBD5E1"),
        top   =Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    num_cols = len(categories) + 2

    # Row 1 – title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    tc = ws.cell(row=1, column=1,
                 value="Maharashtra NEET CAP Round — College-wise Category Cutoff (Closing AIR Rank)")
    tc.font = title_font; tc.fill = title_fill; tc.alignment = center_a
    ws.row_dimensions[1].height = 36

    # Row 2 – sub-header
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sc = ws.cell(row=2, column=1,
                 value=f"MBBS/BDS CAP Round 1  ·  {len(colleges)} Colleges  ·  {len(categories)} Categories")
    sc.font = Font(color="475569", size=10)
    sc.fill = PatternFill("solid", fgColor="F1F5F9")
    sc.alignment = center_a
    ws.row_dimensions[2].height = 22

    # Row 3 – column headers
    c1 = ws.cell(row=3, column=1, value="Code")
    c1.font = h_font; c1.fill = h_fill; c1.alignment = center_a; c1.border = thin
    
    c2 = ws.cell(row=3, column=2, value="College Name")
    c2.font = h_font; c2.fill = h_fill; c2.alignment = left_a; c2.border = thin

    ws.row_dimensions[3].height = 30

    for ci, cat in enumerate(categories, start=3):
        cell = ws.cell(row=3, column=ci, value=cat)
        cell.font = h_font; cell.fill = h_fill
        cell.alignment = center_a; cell.border = thin

    # Rows 4+ – data
    for ri, college in enumerate(colleges, start=4):
        row_fill = even_fill if ri % 2 == 0 else odd_fill
        ws.row_dimensions[ri].height = 18

        cell_code = ws.cell(row=ri, column=1, value=college["college_code"])
        cell_code.font = col_font; cell_code.fill = row_fill
        cell_code.alignment = center_a; cell_code.border = thin

        cell_name = ws.cell(row=ri, column=2, value=college["college_name"])
        cell_name.font = col_font; cell_name.fill = row_fill
        cell_name.alignment = left_a; cell_name.border = thin

        for ci, cat in enumerate(categories, start=3):
            val  = college["category_cutoffs"].get(cat)
            disp = val if val else "—"
            cell = ws.cell(row=ri, column=ci, value=disp)
            cell.font = data_font; cell.fill = row_fill
            cell.alignment = center_a; cell.border = thin

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45
    for ci in range(3, num_cols + 1):  # columns 3..num_cols (= 2 fixed + N categories)
        ws.column_dimensions[get_column_letter(ci)].width = 15

    ws.freeze_panes = "C4"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ─────────────────────────────────────────────────
# JSON helper (handles ObjectId + datetime)
# ─────────────────────────────────────────────────
class _Enc(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, ObjectId): return str(o)
        if isinstance(o, datetime): return o.isoformat()
        return super().default(o)

def _json(data, status=200):
    return app.response_class(
        response=json.dumps(data, cls=_Enc),
        status=status,
        mimetype="application/json",
    )


@app.errorhandler(Exception)
def handle_exception(e):
    if request.path.startswith("/api/"):
        if isinstance(e, HTTPException):
            return _json({"error": e.description}, e.code)
        import traceback
        return _json({"error": str(e), "trace": traceback.format_exc()}, 500)
    if isinstance(e, HTTPException):
        return e
    return "Internal Server Error", 500

# ─────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


tasks = {}

@app.route("/api/upload", methods=["POST"])
def upload_pdf():
    if "file" not in request.files:
        return _json({"error": "No file provided"}, 400)

    file = request.files["file"]
    if not file.filename.lower().endswith(".pdf"):
        return _json({"error": "Only PDF files are accepted"}, 400)

    fname = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{file.filename}"
    fpath = os.path.join(UPLOAD_FOLDER, fname)
    file.save(fpath)

    task_id = str(uuid.uuid4())
    tasks[task_id] = {"status": "processing", "progress": 0, "result": None, "error": None}

    def run_task(fpath, t_id):
        try:
            def prog_cb(curr, tot):
                tasks[t_id]["progress"] = int((curr / tot) * 100)
                
            records = extract_cutoffs(fpath, progress_cb=prog_cb)
            
            if not records:
                tasks[t_id]["status"] = "error"
                tasks[t_id]["error"] = "No cutoff data extracted. Check PDF format."
                return

            all_cats: set = set()
            for r in records:
                all_cats.update(r["category_cutoffs"].keys())

            tasks[t_id]["result"] = {
                "success":          True,
                "doc_id":           None,
                "db_saved":         False,
                "total_colleges":   len(records),
                "total_categories": len(all_cats),
                "categories":       sort_categories(all_cats),
                "records":          records,
            }
            tasks[t_id]["status"] = "done"
            
        except Exception as e:
            import traceback
            tasks[t_id]["status"] = "error"
            tasks[t_id]["error"] = str(e)
            tasks[t_id]["trace"] = traceback.format_exc()
        finally:
            if os.path.exists(fpath):
                try: os.remove(fpath)
                except: pass

    threading.Thread(target=run_task, args=(fpath, task_id)).start()
    return _json({"task_id": task_id, "status": "processing"})

@app.route("/api/upload/status/<task_id>", methods=["GET"])
def upload_status(task_id):
    if task_id not in tasks:
        return _json({"error": "Task not found"}, 404)
    return _json(tasks[task_id])


@app.route("/api/extractions", methods=["GET"])
def list_extractions():
    try:
        docs = list(get_col().find({}, {"records": 0}).sort("upload_time", -1).limit(50))
    except Exception:
        return _json([])
    for d in docs:
        d["_id"] = str(d["_id"])
        if "upload_time" in d:
            d["upload_time"] = d["upload_time"].isoformat()
    return _json(docs)


@app.route("/api/save", methods=["POST"])
def save_extraction():
    data = request.get_json(force=True, silent=True) or {}
    title = data.get("title", "Untitled Document")
    filename = data.get("filename", "unknown.pdf")
    records = data.get("records", [])
    
    if not records:
         return _json({"error": "No records to save"}, 400)
    
    all_cats = set()
    for r in records:
        all_cats.update(r.get("category_cutoffs", {}).keys())
        
    doc = {
        "title": title,
        "filename": filename,
        "upload_time": datetime.utcnow(),
        "records": records,
        "total_colleges": len(records),
        "total_categories": len(all_cats),
    }
    
    try:
        result = get_col().insert_one(doc)
        doc_id = str(result.inserted_id)
        return _json({"success": True, "doc_id": doc_id})
    except Exception as e:
        print(f"[MongoDB] Save failed: {e}")
        return _json({"error": str(e)}, 500)


@app.route("/api/extraction/<doc_id>", methods=["GET"])
def get_extraction(doc_id):
    doc = get_col().find_one({"_id": ObjectId(doc_id)})
    if not doc:
        return _json({"error": "Not found"}, 404)
    doc["_id"] = str(doc["_id"])
    if "upload_time" in doc:
        doc["upload_time"] = doc["upload_time"].isoformat()
        
    all_cats = set()
    for r in doc.get("records", []):
        all_cats.update(r.get("category_cutoffs", {}).keys())
    doc["categories"] = sort_categories(all_cats)
    
    return _json(doc)


@app.route("/api/download/<doc_id>", methods=["GET"])
def download_excel(doc_id):
    doc = get_col().find_one({"_id": ObjectId(doc_id)})
    if not doc:
        return _json({"error": "Not found"}, 404)

    records = doc.get("records", [])
    if not records:
        return _json({"error": "No records"}, 404)

    buf  = generate_excel(records)
    safe = re.sub(r"[^\w\-.]", "_", doc.get("filename", "cutoff"))
    return send_file(
        buf, as_attachment=True,
        download_name=f"{safe}_cutoffs.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/extraction/<doc_id>", methods=["DELETE"])
def delete_extraction(doc_id):
    try:
        get_col().delete_one({"_id": ObjectId(doc_id)})
    except Exception:
        pass
    return _json({"success": True})


@app.route("/api/download-direct", methods=["POST"])
def download_excel_direct():
    """Generate Excel directly from posted records (no DB needed)."""
    data    = request.get_json(force=True, silent=True) or {}
    records = data.get("records", [])
    if not records:
        return _json({"error": "No records provided"}, 400)
    buf = generate_excel(records)
    return send_file(
        buf, as_attachment=True,
        download_name="cutoff_data.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True, port=5000, use_reloader=False)
