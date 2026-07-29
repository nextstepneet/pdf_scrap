import openpyxl
from app.extractor import extract_cutoffs

print("Extracting colleges from PDF using extractor.py...")
recs = extract_cutoffs('SellList-R1-MBBS-BDS.pdf')
extracted_colleges = {str(r['college_code']): str(r['college_name']).strip() for r in recs}
print(f"Total Colleges extracted from PDF: {len(extracted_colleges)}")

print("\n--- Checking cutoff_data.xlsx ---")
wb1 = openpyxl.load_workbook(r'D:\data\cutoff_data.xlsx')
ws1 = wb1.active
excel1_colleges = {}
for i in range(4, ws1.max_row + 1):
    code = ws1.cell(row=i, column=1).value
    name = ws1.cell(row=i, column=2).value
    if code and str(code).strip():
        excel1_colleges[str(code).strip()] = str(name).strip()

print(f"Total Colleges in cutoff_data.xlsx: {len(excel1_colleges)}")
mismatches_1 = []
for code, name in extracted_colleges.items():
    if code not in excel1_colleges:
        mismatches_1.append(f"Code {code} ({name}) not found in Excel.")
    elif excel1_colleges[code] != name:
        mismatches_1.append(f"Code {code}: PDF name '{name}' != Excel name '{excel1_colleges[code]}'")

if not mismatches_1:
    print("Perfect Match! All extracted colleges match cutoff_data.xlsx perfectly.")
else:
    print(f"Found {len(mismatches_1)} mismatches in cutoff_data.xlsx:")
    for m in mismatches_1[:10]:
        print("  -", m)

print("\n--- Checking MBBS_BDS_R1_Cutoffs.xlsx ('All Cutoffs' sheet) ---")
wb2 = openpyxl.load_workbook(r'D:\data\MBBS_BDS_R1_Cutoffs.xlsx')
ws2 = wb2['All Cutoffs']
excel2_colleges = {}
for i in range(2, ws2.max_row + 1):
    code = ws2.cell(row=i, column=1).value
    name = ws2.cell(row=i, column=2).value
    if code and str(code).strip():
        excel2_colleges[str(code).strip()] = str(name).strip()

print(f"Total Colleges in MBBS_BDS_R1_Cutoffs.xlsx: {len(excel2_colleges)}")
mismatches_2 = []
for code, name in extracted_colleges.items():
    if code not in excel2_colleges:
        mismatches_2.append(f"Code {code} ({name}) not found in Excel.")
    elif excel2_colleges[code] != name:
        mismatches_2.append(f"Code {code}: PDF name '{name}' != Excel name '{excel2_colleges[code]}'")

if not mismatches_2:
    print("Perfect Match! All extracted colleges match MBBS_BDS_R1_Cutoffs.xlsx perfectly.")
else:
    print(f"Found {len(mismatches_2)} mismatches in MBBS_BDS_R1_Cutoffs.xlsx:")
    for m in mismatches_2[:10]:
        print("  -", m)

